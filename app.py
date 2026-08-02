"""
===============================================================================
AI THERMAL TOMATO ANALYZER - WEB BACKEND SERVER
===============================================================================
Đơn vị phát triển : LeafOptiAI Research Team
Trường             : Hanoi Pedagogical University 2 (HPU2)
===============================================================================
"""

import os
import sys
import time
import uuid
import logging
import threading
from datetime import datetime
from typing import Tuple, Dict, Any, List
from urllib.parse import urlparse, unquote

from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    send_from_directory,
    url_for
)
# Import bộ phân tích AI thực tế của bạn
from models.predict import predict_thermal_image, Config # (Thay đổi tên hàm/class cho khớp với code thật của bạn)

import cv2
import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    import csv
    HAS_OPENPYXL = False

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image as RLImage,
    HRFlowable,
    KeepTogether
)
from reportlab.pdfgen import canvas


class Config:
    SECRET_KEY = os.environ.get('SECRET_KEY', 'leafoptiai-hpu2-thermal-secret-key-2026')
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    
    # Cấu hình đầy đủ cây thư mục dự án
    UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
    RESULTS_FOLDER = os.path.join(BASE_DIR, 'results')
    ORIGINAL_FOLDER = os.path.join(RESULTS_FOLDER, 'original')
    SEGMENTATION_FOLDER = os.path.join(RESULTS_FOLDER, 'segmentation')
    BINARY_FOLDER = os.path.join(RESULTS_FOLDER, 'binary')
    
    PDF_FOLDER = os.path.join(BASE_DIR, 'pdf')
    EXCEL_FOLDER = os.path.join(BASE_DIR, 'exports')
    MODEL_PATH = os.path.join(BASE_DIR, 'models', 'best.pt')
    
    MAX_CONTENT_LENGTH = 32 * 1024 * 1024
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'tif', 'tiff', 'bmp'}
    MIN_CONTOUR_AREA = 30
    AUTO_CLEANUP_HOURS = 2


logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s in %(module)s: %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger('ThermalAnalyzer')

app = Flask(__name__)
app.config.from_object(Config)

# Tự động tạo tất cả các thư mục cần thiết
ALL_PROJECT_FOLDERS = [
    Config.UPLOAD_FOLDER,
    Config.RESULTS_FOLDER,
    Config.ORIGINAL_FOLDER,
    Config.SEGMENTATION_FOLDER,
    Config.BINARY_FOLDER,
    Config.PDF_FOLDER,
    Config.EXCEL_FOLDER
]

for directory in ALL_PROJECT_FOLDERS:
    os.makedirs(directory, exist_ok=True)


def is_allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in Config.ALLOWED_EXTENSIONS


def evaluate_canopy_status(coverage_pct: float) -> Dict[str, str]:
    if coverage_pct < 30.0:
        return {"status": "Low Canopy", "description": "Mật độ tán lá thấp.", "color": "#dc3545"}
    elif 30.0 <= coverage_pct <= 65.0:
        return {"status": "Moderate Canopy", "description": "Tán lá phát triển ổn định.", "color": "#28a745"}
    else:
        return {"status": "Dense Canopy", "description": "Tán lá rất dày đặc.", "color": "#155724"}


def cleanup_old_files():
    now = time.time()
    cutoff = now - (Config.AUTO_CLEANUP_HOURS * 3600)
    for folder in ALL_PROJECT_FOLDERS:
        if os.path.exists(folder):
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                if os.path.isfile(file_path) and os.path.getmtime(file_path) < cutoff:
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        logger.error(f"Error removing file {file_path}: {e}")


def start_background_cleanup():
    def run_loop():
        while True:
            time.sleep(1800)
            cleanup_old_files()

    thread = threading.Thread(target=run_loop, daemon=True)
    thread.start()

start_background_cleanup()


class ThermalImageAnalyzer:
    def __init__(self, model_path: str):
        self.model_path = model_path

    def process(self, image_path: str) -> Tuple[np.ndarray, np.ndarray, Dict[str, Any]]:
        start_time = time.time()
        
        original_bgr = cv2.imread(image_path)
        if original_bgr is None:
            raise ValueError("Không thể đọc file ảnh.")
            
        h, w = original_bgr.shape[:2]
        total_pixels = h * w
        
        try:
            from models.predict import predict_thermal_image
            seg_bgr, binary_mask, custom_metrics = predict_thermal_image(image_path, self.model_path)
            
            inference_time = round((time.time() - start_time) * 1000, 1)
            custom_metrics['inference_time'] = inference_time
            custom_metrics['canopy_status'] = evaluate_canopy_status(custom_metrics.get('coverage', 0.0))
            return seg_bgr, binary_mask, custom_metrics

        except (ImportError, Exception) as e:
            logger.info(f"Fallback OpenCV Processor: {e}")
            
            gray = cv2.cvtColor(original_bgr, cv2.COLOR_BGR2GRAY)
            filtered = cv2.bilateralFilter(gray, 9, 75, 75)
            _, binary_mask = cv2.threshold(filtered, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
            
            kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
            binary_mask = cv2.morphologyEx(binary_mask, cv2.MORPH_CLOSE, kernel, iterations=2)
            
            contours, _ = cv2.findContours(binary_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            valid_contours = [cnt for cnt in contours if cv2.contourArea(cnt) >= Config.MIN_CONTOUR_AREA]
            
            seg_bgr = original_bgr.copy()
            overlay = original_bgr.copy()
            cv2.drawContours(overlay, valid_contours, -1, (36, 179, 0), -1)
            cv2.addWeighted(overlay, 0.45, seg_bgr, 0.55, 0, seg_bgr)
            cv2.drawContours(seg_bgr, valid_contours, -1, (0, 255, 64), 2)
            
            mask_leaf_pixels = int(np.count_nonzero(binary_mask))
            coverage = round((mask_leaf_pixels / total_pixels) * 100, 1)
            detected_regions = len(valid_contours)
            mean_conf = round(92.5 if detected_regions > 0 else 0.0, 1)
            inference_time = round((time.time() - start_time) * 1000, 1)
            
            metrics = {
                "leaf_area": mask_leaf_pixels,
                "coverage": coverage,
                "detected_regions": detected_regions,
                "confidence": mean_conf,
                "inference_time": inference_time,
                "canopy_status": evaluate_canopy_status(coverage)
            }
            return seg_bgr, binary_mask, metrics


analyzer = ThermalImageAnalyzer(Config.MODEL_PATH)


# ===============================================================================
# REPORTLAB PDF GENERATOR (CHUẨN 100% GIAO DIỆN MẪU ẢNH 1)
# ===============================================================================
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#444444"))
        self.setStrokeColor(colors.HexColor("#1b5e20"))
        self.setLineWidth(1)
        self.line(40, 40, 555, 40)
        self.drawString(40, 25, "AI Thermal Tomato Analyzer — LeafOptiAI Research Team (HPU2)")
        self.drawRightString(555, 25, f"Page {self._pageNumber} of {page_count}")
        self.restoreState()


def create_pdf_report(output_pdf_path: str, metrics: Dict[str, Any], orig_img_path: str, seg_img_path: str, mask_img_path: str) -> bool:
    try:
        doc = SimpleDocTemplate(
            output_pdf_path, 
            pagesize=A4, 
            leftMargin=40, 
            rightMargin=40, 
            topMargin=40, 
            bottomMargin=50
        )
        styles = getSampleStyleSheet()
        PRIMARY_COLOR = colors.HexColor("#1b5e20")
        
        title_style = ParagraphStyle(
            'DocTitle', 
            parent=styles['Heading1'], 
            fontName='Helvetica-Bold', 
            fontSize=18, 
            textColor=PRIMARY_COLOR, 
            alignment=1, 
            spaceAfter=3
        )
        sub_style = ParagraphStyle(
            'DocSub', 
            parent=styles['Normal'], 
            fontName='Helvetica-Oblique', 
            fontSize=9.5, 
            textColor=colors.HexColor("#333333"), 
            alignment=1, 
            spaceAfter=14
        )
        sec_style = ParagraphStyle(
            'SecTitle', 
            parent=styles['Heading2'], 
            fontName='Helvetica-Bold', 
            fontSize=11, 
            textColor=PRIMARY_COLOR, 
            spaceBefore=14, 
            spaceAfter=8
        )
        text_style = ParagraphStyle('BodyTextCustom', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=12)
        bold_style = ParagraphStyle('BodyBoldCustom', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=12)
        header_table_style = ParagraphStyle('HeaderTableStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.white)

        story = []
        
        story.append(Paragraph("AI Thermal Tomato Analyzer", title_style))
        story.append(Paragraph("Deep Learning-Based Thermal Image Analysis for Tomato Canopy", sub_style))
        story.append(HRFlowable(width="100%", thickness=1.5, color=PRIMARY_COLOR, spaceAfter=14))

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        story.append(Paragraph(f"<b>Export Timestamp:</b> {now_str}", text_style))
        story.append(Spacer(1, 10))

        story.append(Paragraph("I. Quantitative Analysis Metrics", sec_style))
        
        leaf_area = int(metrics.get('leaf_area', 0))
        coverage = float(metrics.get('coverage', 0.0))
        detected_regions = metrics.get('detected_regions', 0)
        
        conf = float(metrics.get('confidence', 0.0))
        conf_pct = conf * 100 if conf <= 1.0 else conf
        
        inference_time = float(metrics.get('inference_time', 0.0))

        table_data = [
            [Paragraph("Parameter Metric", header_table_style), Paragraph("Measured Value", header_table_style), Paragraph("Unit", header_table_style)],
            [Paragraph("Leaf Area", text_style), Paragraph(f"{leaf_area:,}", text_style), Paragraph("pixels", text_style)],
            [Paragraph("Canopy Coverage", text_style), Paragraph(f"{coverage:.1f}%", text_style), Paragraph("%", text_style)],
            [Paragraph("Detected Regions", text_style), Paragraph(str(detected_regions), text_style), Paragraph("count", text_style)],
            [Paragraph("Mean Confidence", text_style), Paragraph(f"{conf_pct:.1f}%", text_style), Paragraph("%", text_style)],
            [Paragraph("Inference Time", text_style), Paragraph(f"{inference_time:.1f} ms", text_style), Paragraph("ms", text_style)],
        ]
        
        t = Table(table_data, colWidths=[220, 195, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PRIMARY_COLOR),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#d3d3d3")),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#fcfcfc")]),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(t)
        story.append(Spacer(1, 15))

        story.append(Paragraph("II. Thermal Image Visualizations", sec_style))
        img_w, img_h = 2.2 * 72, 1.7 * 72
        
        def load_img_safely(path):
            if path and os.path.isfile(path) and os.path.getsize(path) > 0:
                try:
                    return RLImage(path, width=img_w, height=img_h)
                except Exception as e:
                    logger.error(f"Error loading image into PDF: {e}")
            return Paragraph("<i>Image not found</i>", text_style)

        img_orig = load_img_safely(orig_img_path)
        img_seg = load_img_safely(seg_img_path)
        img_mask = load_img_safely(mask_img_path)
        
        img_table_data = [
            [img_orig, img_seg, img_mask],
            [Paragraph("Original Image", bold_style), Paragraph("Segmentation Result", bold_style), Paragraph("Binary Mask", bold_style)]
        ]
        
        img_table = Table(img_table_data, colWidths=[171, 171, 171])
        img_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
        ]))
        story.append(KeepTogether([img_table]))
        
        doc.build(story, canvasmaker=NumberedCanvas)
        return True
    except Exception as e:
        logger.error(f"Error generating PDF: {e}", exc_info=True)
        return False


# ===============================================================================
# EXCEL EXPORTER
# ===============================================================================
def create_excel_report(output_path: str, history_items: List[Dict[str, Any]]) -> bool:
    try:
        if HAS_OPENPYXL:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Analysis History"

            ws.merge_cells("A1:G1")
            ws["A1"] = "BÁO CÁO LỊCH SỬ PHÂN TÍCH LÁ CÀ CHUA (LEAFOPTI AI)"
            ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="1B5E20")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A2:G2")
            ws["A2"] = f"Thời gian xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            ws["A2"].font = Font(name="Segoe UI", size=9, italic=True, color="555555")
            ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

            headers = ["STT", "Tên File Ảnh", "Thời Gian", "Diện Tích Lá (px)", "Độ Phủ (%)", "Độ Tin Cậy (%)", "Thời Gian Xử Lý (ms)"]
            
            header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
            header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

            for col_num, header_title in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num, value=header_title)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            row_idx = 5
            for idx, item in enumerate(history_items, 1):
                conf = item.get('confidence', 0)
                conf_pct = round(conf * 100 if conf <= 1.0 else conf, 1)

                row_values = [
                    idx,
                    item.get('filename', 'N/A'),
                    item.get('timestamp', datetime.now().strftime('%H:%M:%S %d/%m/%Y')),
                    item.get('leaf_area', 0),
                    item.get('coverage', 0.0),
                    conf_pct,
                    item.get('inference_time', 0.0)
                ]

                for col_num, val in enumerate(row_values, 1):
                    cell = ws.cell(row=row_idx, column=col_num, value=val)
                    cell.font = Font(name="Segoe UI", size=9.5)
                    cell.border = thin_border
                    
                    if col_num in [1, 3, 4, 5, 6, 7]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                row_idx += 1

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[4].height = 24

            wb.save(output_path)
            return True
        else:
            with open(output_path, mode='w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(["STT", "Tên File Ảnh", "Thời Gian", "Diện Tích Lá (px)", "Độ Phủ (%)", "Độ Tin Cậy (%)", "Thời Gian Xử Lý (ms)"])
                for idx, item in enumerate(history_items, 1):
                    conf = item.get('confidence', 0)
                    conf_pct = round(conf * 100 if conf <= 1.0 else conf, 1)
                    writer.writerow([
                        idx,
                        item.get('filename', 'N/A'),
                        item.get('timestamp', datetime.now().strftime('%H:%M:%S %d/%m/%Y')),
                        item.get('leaf_area', 0),
                        item.get('coverage', 0.0),
                        conf_pct,
                        item.get('inference_time', 0.0)
                    ])
            return True
    except Exception as e:
        logger.error(f"Lỗi tạo Excel: {e}", exc_info=True)
        return False


# ===============================================================================
# HÀM ĐỊNH VỊ ẢNH THÔNG MINH - TÌM TẤT CẢ THƯ MỤC CON
# ===============================================================================
def locate_image_file(raw_input: Any, folder_keywords: List[str], file_keywords: List[str]) -> str:
    """
    1. Nếu có tên file: Quét toàn bộ thư mục dự án (BASE_DIR) để tìm chính xác file.
    2. Nếu không tìm thấy: Tìm file ảnh mới nhất theo từ khóa trong tất cả thư mục.
    """
    if raw_input:
        path_str = unquote(urlparse(str(raw_input)).path)
        clean_name = os.path.basename(path_str)
        if clean_name:
            for root, _, files in os.walk(Config.BASE_DIR):
                if clean_name in files:
                    full_p = os.path.join(root, clean_name)
                    if os.path.isfile(full_p) and os.path.getsize(full_p) > 0:
                        return full_p

    # Dự phòng: Duyệt các thư mục ảnh chuẩn
    search_dirs = [
        Config.ORIGINAL_FOLDER,
        Config.SEGMENTATION_FOLDER,
        Config.BINARY_FOLDER,
        Config.UPLOAD_FOLDER,
        Config.RESULTS_FOLDER
    ]
    
    candidates = []
    for s_dir in search_dirs:
        if os.path.exists(s_dir):
            for fname in os.listdir(s_dir):
                fpath = os.path.join(s_dir, fname)
                if os.path.isfile(fpath) and os.path.getsize(fpath) > 0:
                    fname_lower = fname.lower()
                    s_dir_lower = s_dir.lower()
                    if any(kw in fname_lower for kw in file_keywords) or any(kw in s_dir_lower for kw in folder_keywords):
                        candidates.append(fpath)

    if candidates:
        candidates.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        return candidates[0]

    return ''


# ===============================================================================
# ROUTES / API ENDPOINTS
# ===============================================================================
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/analyze', methods=['POST'])
def analyze():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Không có file gửi lên.'}), 400
    
    file = request.files['file']
    if file.filename == '' or not is_allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'File không hợp lệ.'}), 400

    try:
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'png'
        token = uuid.uuid4().hex[:8]
        
        orig_name = f"thermal_orig_{token}.{ext}"
        seg_name = f"thermal_seg_{token}.png"
        mask_name = f"thermal_mask_{token}.png"
        
        # Đường dẫn lưu CHỈ vào các thư mục con tương ứng
        orig_path = os.path.join(app.config['ORIGINAL_FOLDER'], orig_name)
        seg_path = os.path.join(app.config['SEGMENTATION_FOLDER'], seg_name)
        mask_path = os.path.join(app.config['BINARY_FOLDER'], mask_name)
        
        # Lưu file vào thư mục con chuẩn
        file.save(orig_path)
        seg_bgr, binary_mask, metrics = analyzer.process(orig_path)
        
        cv2.imwrite(seg_path, seg_bgr)
        cv2.imwrite(mask_path, binary_mask)
        
        # Đã loại bỏ các lệnh ghi đè dư thừa ra ngoài thư mục results/ và uploads/
        
        return jsonify({
            'success': True,
            'images': {
                'original': url_for('get_result_file', filename=f"original/{orig_name}"),
                'segmentation': url_for('get_result_file', filename=f"segmentation/{seg_name}"),
                'binary_mask': url_for('get_result_file', filename=f"binary/{mask_name}")
            },
            'metrics': metrics,
            'filenames': {
                'original': orig_name,
                'segmentation': seg_name,
                'binary_mask': mask_name,
                'raw_name': file.filename
            }
        }), 200

    except Exception as e:
        logger.error(f"Lỗi /analyze: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/export_pdf', methods=['POST'])
def export_pdf():
    try:
        data = request.get_json() or {}
        
        metrics = data.get('metrics', {})
        filenames = data.get('filenames', {})
        images = data.get('images', {})

        orig_raw, seg_raw, mask_raw = '', '', ''

        if isinstance(filenames, dict):
            orig_raw = filenames.get('original', '')
            seg_raw = filenames.get('segmentation', '')
            mask_raw = filenames.get('binary_mask', '')
        elif isinstance(filenames, list) and len(filenames) > 0:
            if isinstance(filenames[0], dict):
                orig_raw = filenames[0].get('original', '')
                seg_raw = filenames[0].get('segmentation', '')
                mask_raw = filenames[0].get('binary_mask', '')
            elif isinstance(filenames[0], str):
                orig_raw = filenames[0] if len(filenames) > 0 else ''
                seg_raw = filenames[1] if len(filenames) > 1 else ''
                mask_raw = filenames[2] if len(filenames) > 2 else ''

        if not orig_raw and isinstance(images, dict):
            orig_raw = images.get('original', '')
            seg_raw = images.get('segmentation', '')
            mask_raw = images.get('binary_mask', '')

        # Tìm kiếm chính xác trong tất cả thư mục dự án
        orig_path = locate_image_file(orig_raw, ['original', 'upload'], ['orig', 'upload'])
        seg_path = locate_image_file(seg_raw, ['segmentation'], ['seg'])
        mask_path = locate_image_file(mask_raw, ['binary'], ['mask', 'bin'])

        pdf_filename = f"AI_Thermal_Report_{uuid.uuid4().hex[:6]}.pdf"
        pdf_out_path = os.path.join(app.config['PDF_FOLDER'], pdf_filename)

        if create_pdf_report(pdf_out_path, metrics, orig_path, seg_path, mask_path):
            return jsonify({
                'success': True,
                'pdf_url': url_for('download_pdf', filename=pdf_filename),
                'filename': pdf_filename
            }), 200
        
        return jsonify({'success': False, 'error': 'Lỗi trong quá trình tạo file PDF.'}), 500

    except Exception as e:
        logger.error(f"Lỗi route /export_pdf: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/export_excel', methods=['POST'])
def export_excel():
    data = request.get_json() or {}
    history_data = data.get('history', [])
    
    if not history_data and 'metrics' in data:
        m = data['metrics']
        fns = data.get('filenames', {})
        
        fname_raw = 'thermal_image.jpg'
        if isinstance(fns, dict):
            fname_raw = fns.get('raw_name', fns.get('original', fname_raw))
        elif isinstance(fns, list) and len(fns) > 0 and isinstance(fns[0], dict):
            fname_raw = fns[0].get('raw_name', fns[0].get('original', fname_raw))

        history_data = [{
            'filename': fname_raw,
            'timestamp': datetime.now().strftime('%H:%M:%S %d/%m/%Y'),
            'leaf_area': m.get('leaf_area', 0),
            'coverage': m.get('coverage', 0.0),
            'confidence': m.get('confidence', 0.0),
            'inference_time': m.get('inference_time', 0.0)
        }]

    if not history_data:
        return jsonify({'success': False, 'error': 'Chưa có dữ liệu lịch sử để xuất Excel.'}), 400

    ext = "xlsx" if HAS_OPENPYXL else "csv"
    excel_filename = f"LeafOptiAI_Analysis_History_{uuid.uuid4().hex[:6]}.{ext}"
    excel_out_path = os.path.join(app.config['EXCEL_FOLDER'], excel_filename)

    if create_excel_report(excel_out_path, history_data):
        return jsonify({
            'success': True,
            'excel_url': url_for('download_excel', filename=excel_filename),
            'filename': excel_filename
        }), 200

    return jsonify({'success': False, 'error': 'Lỗi trong quá trình tạo file Excel.'}), 500


@app.route('/uploads/<path:filename>')
def get_upload_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route('/results/<path:filename>')
def get_result_file(filename):
    return send_from_directory(app.config['RESULTS_FOLDER'], filename)


@app.route('/pdf/<path:filename>')
def download_pdf(filename):
    return send_from_directory(app.config['PDF_FOLDER'], filename, as_attachment=True)


@app.route('/excel/<path:filename>')
def download_excel(filename):
    return send_from_directory(app.config['EXCEL_FOLDER'], filename, as_attachment=True)


if __name__ == '__main__':
    logger.info("Khởi chạy Server Flask trên http://127.0.0.1:5000")
    app.run(host='0.0.0.0', port=5000, debug=True)
